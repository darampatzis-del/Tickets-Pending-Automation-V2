import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font
from datetime import datetime
from copy import copy
import shutil, sys, os, re, subprocess
import numpy as np


#> TEXT FILES
def read_txt_file(file_name):
    try:
        with open(file_name, "r") as f:
            return_list = []
            values = list(map(str.rstrip, f))
            for value in values:
                value = value.split(';')
                return_list.append(value)
            return return_list
    except FileNotFoundError:
        print(f"The file '{file_name}' not found.")
        sys.exit(1)
    except:
        print(f"An error occurred while reading the file '{file_name}'.")
        sys.exit(1)
        
def format_list(list, value):
    return_str = ""
    if list is not None:
        for element in list:
            # Create a regex pattern to match whole words from the customers and queues list
            # Gets list as argument but processes only the first element as we split the list with delimiter ";". for instance "FI-CO;fi". We only care about FI-CO
            pattern = r"\b" + re.escape(element[0]) + r"\b"
            if re.search(pattern, value):
                return_str = "".join(element[0])
    return return_str
#< TEXT FILES


#> FILE HANDLING
def check_if_file_closed(file_path):
    try:
        # Try opening the file in read mode
        with open(file_path, "r"):
            return True  # File is closed and accessible
    except (PermissionError, OSError):
        # File is either open or doesn't exist
        return False


def pd_open_excel_file(file, sheet_name):
    if os.path.exists(file):
        try:
            if check_if_file_closed(file):
                df = pd.read_excel(file, sheet_name=sheet_name, engine="openpyxl")
                return df
            else:
                print(f"File '{file}' is still open or inaccessible. Please close it.")
        except Exception as e:
            print(f"Cannot open excel file {file} with error: {e}")
            sys.exit(1)
    else:
        print(f"File {file} does not exist")
        sys.exit(1)


def wb_load_workbook(file):
    if os.path.exists(file):
        if check_if_file_closed(file):
            try:
                wb = load_workbook(file, read_only=False, data_only=True)
                return wb
            except Exception as e:
                print(f"An error occurred while loading the file: {e}")
        else:
            print(f"File '{file}' is still open or inaccessible. Please close it.")
    else:
        print(f"File {file} does not exist")
        sys.exit(1)
#< FILE HANDLING


#> PRIORITY HANDLING
def sort_priority(df):
    df["Subject"] = df["Subject"].str.strip()
    regex_pattern = r"prior|urg|alt|high"
    matching_rows = df["Subject"].str.contains(
        regex_pattern, case=False, regex=True, na=False
    )
    df.loc[matching_rows, "Priority"] = "2 High"

    df["Created"] = pd.to_datetime(
        df["Created"], errors="coerce"
    )  # change 'Created' to datetime and on error ignore the cell
    df = df.sort_values(by=["Priority", "Created"], ascending=[True, False])

    # Convert "Ticket Number" to string because excel displays it as scientific
    df["Ticket Number"] = df["Ticket Number"].astype("string")

    return df


def color_priority(file):
    wb = wb_load_workbook(file)
    sheet = wb.active

    ignore_header = 2
    len_columns = sheet.max_column

    for row in sheet.iter_rows(min_row=ignore_header, values_only=False):
        for cell in row:
            if cell.value == "2 High":
                for col_idx in range(1, len_columns + 1):
                    cell = sheet.cell(row=cell.row, column=col_idx)
                    cell.font = Font(color="FF0000")
    wb.save(file)
#< PRIORITY HANDLING


#> COPY ROWS
def copy_rows(file, sheet_list):
    wb = wb_load_workbook(file)

    # Create target sheet if not exists
    if sheet_list["name"] not in wb.sheetnames:
        print(f"Sheet '{sheet_list['name']}' missing. Creating it.")
        wb.create_sheet(sheet_list["name"])

    source_sheet = wb["All"]
    target_sheet = wb[sheet_list["name"]]

    # Find the index of the "Queue" column
    queue_column_index = None
    for col_index, cell in enumerate(source_sheet[1], 1):  # source_sheet[1] = second row, 1 = start counting from 1 and not 0
        if cell.value == "Queue":
            queue_column_index = col_index
            break

    if queue_column_index is None:
        print(f"Column 'Queue' not found in sheet '{sheet_list['name']}'.")
        return

    target_row_index = 2  # Start from 2nd row (header is 1st)

    # Iterate through rows starting from row 2 (ignoring header)
    for row in source_sheet.iter_rows(min_row=2, values_only=False):  # Iterate through all data rows
        queue_value = row[queue_column_index - 1].value

        # If the value in the "Queue" column matches one in sheet_list['search'], copy the entire row
        if queue_value in sheet_list["search"]:
            for col_idx, cell in enumerate(row, start=1):
                target_cell = target_sheet.cell(row=target_row_index, column=col_idx)
                target_cell.value = copy(clean_string(cell.value))  # Copy the value
                target_cell.style = copy(cell.style)  # Copy the style
                target_cell.font = copy(cell.font)  # Copy the font style
                if isinstance(clean_string(cell.value), datetime):
                    target_cell.number_format = (cell.number_format)  # Copy number format for datetime values

            target_row_index += 1  # Move to the next row for the next match

    # Save the workbook after making changes
    wb.save(file)
#< COPY ROWS


#> ATTEMPT TO CLEAR ERROR WHEN EXCEL OPENS.
def clean_string(val):
    if isinstance(val, str):
        # Remove non-printable and invalid XML characters
        return re.sub(r"[\x00-\x08\x0B\x0C\x0E-\x1F\x7F-\x9F]", "", val)
    return val
#< ATTEMPT TO CLEAR ERROR WHEN EXCEL OPENS.


def prepare_data(source_file, columns_to_copy, sheet_name="Sheet1"):
    df = pd_open_excel_file(source_file, sheet_name)

    # Extract the values from the "Queue" column
    queue_col = df["Queue"]

    # Create a list to hold the transformed data
    transformed_data = []

    customers = read_txt_file("Customers.txt")
    queues = read_txt_file("Queues.txt")
    
    # Process each value in the "queue" column
    for value in queue_col:
        customer_str = format_list(customers, value)
        queue_str = format_list(queues, value)
        if not queue_str.strip():
            queue_str = customer_str
        transformed_data.append([customer_str, queue_str])

    # Convert transformed_data to DataFrame
    transformed_data = pd.DataFrame(transformed_data, columns=["Customer", "Queue"])

    # Join Customer and Queue columns with selected columns
    df_selected_columns = pd.DataFrame(df[columns_to_copy])

    # If NaN values are in the sheet, replace with empty string
    df = transformed_data.join(df_selected_columns).fillna("")

    df = df.map(lambda x: x.replace("\n", " ").replace("\r", "") if isinstance(x, str) else x)
    df = df.map(lambda x: re.sub(r"[\r\n\x00-\x1f\x7f-\x9f]", " ", x) if isinstance(x, str) else x)
    df.columns = df.columns.str.replace(r"[\[\]\*\/\\\?\:]", "", regex=True)

    # Trim all strings
    df = df.map(lambda x: x.strip() if isinstance(x, str) else x)

    df = df.map(clean_string)
    df = df.replace({np.nan: None})
    if pd.isna(value):
        value = None

    # Sort by priority
    df = sort_priority(df)

    return df


def main(source_file):
    # Generate file with current date for instance "Tickets Pending 31-02-2026.xlsx"
    current_date = datetime.today().strftime("%d-%m-%Y")
    template_file = "Template.xlsx"
    target_file = f"Tickets Pending {current_date}.xlsx"

    # Copy the template to create a new file
    shutil.copy(template_file, target_file)

    sheets = {
        "all": "All",
        "SD_CS": {
            "name": "SD_CS",
            "search": [],
            "copy": True
        },
        "MM_PP_QM": {
            "name": "MM_PP_QM",
            "search": [],
            "copy": True,
        },
        "FI-CO": {
            "name": "FI-CO",
            "search": [],
            "copy": True
        },
        "System": {
            "name": "System",
            "search": [],
            "copy": True,
        }
    }

    header_list = [
        "Ticket Number",
        "Subject",
        "Age",
        "Created",
        "Priority",
        "CustomerID",
        "Customer Name",
        "From",
        "Type"
    ]

    try:
        data = prepare_data(source_file, header_list)
        target_wb = load_workbook(target_file)
        sheet = target_wb[sheets["all"]]
        
        # Populate sheets "search" with the correct strings
        queues = read_txt_file("Queues.txt")
    
        for queue in queues:
            if len(queue) > 1 and queue[1] in sheets:
                sheets[queue[1]]["search"].append(queue[0])
                
        next_row = sheet.max_row + 1  # Start appending from the next available row

        # Append data from the DataFrame to the Excel sheet
        for _, row in data.iterrows():
            for col_idx, value in enumerate(row, start=1):  # start=1 to begin from column 1
                cell = sheet.cell(row=next_row, column=col_idx, value=value)
                if isinstance(value, datetime):
                    cell.number_format = "yyyy-mm-dd hh:mm"
                elif isinstance(value, float) or isinstance(value, int):
                    cell.number_format = "General"
                if pd.isna(value) or (isinstance(value, str) and value.strip().lower() == "nan"):
                    cell.value = None
                else:
                    cell.value = clean_string(value)
            next_row += 1

        target_wb.save(target_file)
        color_priority(target_file)

        for key, value in sheets.items():
            if isinstance(value, dict) and value.get("copy") == True:
                copy_rows(target_file, sheets[key])
        
    except Exception as e:
        print(f"An error occurred during processing: {e}")

    print(f"New Excel file created: {target_file}")


if __name__ == "__main__":
    import sys
    # Argument: raw file from OTRS
    main(sys.argv[1])